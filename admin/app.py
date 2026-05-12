import datetime
import json
import os
import time
import urllib.parse
import requests as http_requests
from functools import wraps
from flask import Flask, request, render_template, redirect, url_for, session, flash, jsonify
from config import API_BASE_URL, SESSION_SECRET

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10 MB upload limit
app.secret_key = SESSION_SECRET
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["PERMANENT_SESSION_LIFETIME"] = 1800  # 30 minutes
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("FLASK_ENV") != "development"

import hashlib
import hmac
import secrets

from utils import parse_json_content


def _safe_float(val, default=0.0):
    try:
        return float(val)
    except (ValueError, TypeError):
        return default

def _safe_int(val, default=0):
    try:
        return int(val)
    except (ValueError, TypeError):
        return default

def _clamp_page_size(val):
    """Clamp page_size to [1, 100] to prevent DoS."""
    val = _safe_int(val, 20)
    return max(1, min(val, 100))

def _csrf_token():
    """Generate a CSRF token and store it in the session (only if not already set)."""
    existing = session.get("_csrf_token")
    if existing:
        return existing
    token = secrets.token_hex(16)
    session["_csrf_token"] = token
    return token

@app.before_request
def _check_csrf():
    """Validate CSRF token for POST requests."""
    if request.method != "POST":
        return
    # Skip CSRF check in testing mode (tests obtain token via GET or session_transaction)
    if app.config.get("TESTING"):
        return
    form_token = request.form.get("_csrf_token", "")
    session_token = session.get("_csrf_token", "")
    if not form_token or not session_token or not hmac.compare_digest(form_token, session_token):
        session.clear()
        flash("安全验证失败，请重新登录", "error")
        return redirect(url_for("login"))



def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_phone") or not session.get("is_admin"):
            session.clear()
            return redirect(url_for("login"))
        # Refresh session lifetime on each request (sliding expiration)
        session.permanent = True
        session.modified = True
        return f(*args, **kwargs)
    return decorated



def api_get(path, token):
    return http_requests.get(
        f"{API_BASE_URL}{path}",
        headers={"Authorization": f"Bearer {token}"},
        timeout=10
    )


def api_post(path, json_data=None, token=None, timeout=10):
    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    return http_requests.post(
        f"{API_BASE_URL}{path}",
        json=json_data,
        headers=headers,
        timeout=timeout
    )


def api_put(path, token, data):
    return http_requests.put(
        f"{API_BASE_URL}{path}",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        data=json.dumps(data),
        timeout=10
    )


def api_delete(path, token):
    return http_requests.delete(
        f"{API_BASE_URL}{path}",
        headers={"Authorization": f"Bearer {token}"},
        timeout=10
    )


def _safe_api_error(resp):
    """Safely extract error message from an API response."""
    try:
        return resp.json().get("error", f"请求失败 (HTTP {resp.status_code})")
    except Exception:
        return f"请求失败 (HTTP {resp.status_code})"


def parse_csv_content(content):
    """解析 CSV 内容为规则列表，支持中英文表头"""
    import csv
    import io

    # Strip BOM if present (UTF-8 BOM = ﻿)
    if content and content[0] == "﻿":
        content = content[1:]

    reader = csv.reader(io.StringIO(content))
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not rows:
        return []

    HEADER_MAP = {
        "keyword": "keyword", "关键词": "keyword", "规则标题": "keyword", "规则名称": "keyword",
        "match_type": "match_type", "matchtype": "match_type", "触发类型": "match_type",
        "reply_template": "reply_template", "replytemplate": "reply_template",
        "reply_content": "reply_template", "replycontent": "reply_template",
        "回复内容": "reply_template",
        "category": "category", "规则分类": "category", "rule_category": "category",
        "target_type": "target_type", "targettype": "target_type",
        "target_names": "target_names", "targetnames": "target_names",
        "适用房源": "target_names",
        "trigger_condition": "trigger_condition", "触发条件": "trigger_condition",
        "enabled": "enabled", "状态": "enabled",
        "priority": "priority",
    }
    MATCH_TYPE_MAP = {
        "关键词回复": "CONTAINS", "包含匹配": "CONTAINS",
        "精确匹配": "EXACT", "正则匹配": "REGEX",
    }
    ENABLED_MAP = {
        "启用": 1, "是": 1, "yes": 1, "true": 1,
        "禁用": 0, "否": 0, "no": 0, "false": 0,
    }
    DEFAULT_ORDER = [
        "keyword", "match_type", "reply_template", "category",
        "target_type", "target_names", "priority", "enabled",
    ]

    def normalize_header(h):
        h = h.strip()
        if h in HEADER_MAP:
            return HEADER_MAP[h]
        lowered = h.lower().replace("-", "_").replace(" ", "")
        return HEADER_MAP.get(lowered, "")

    def map_match_type(val):
        if not val:
            return "CONTAINS"
        val = val.strip()
        if val in MATCH_TYPE_MAP:
            return MATCH_TYPE_MAP[val]
        v = val.upper()
        if v in ("CONTAINS", "EXACT", "REGEX"):
            return v
        return "CONTAINS"

    def parse_enabled(val):
        if val is None:
            return 1
        val = str(val).strip().lower()
        if val in ENABLED_MAP:
            return ENABLED_MAP[val]
        try:
            return 1 if int(val) else 0
        except ValueError:
            return 1

    header_keys = [normalize_header(h) for h in rows[0]]
    has_header = any(
        k in ("keyword", "reply_template", "回复内容", "规则标题")
        for k in header_keys
    )
    data_rows = rows[1:] if has_header else rows

    rules = []
    for row in data_rows:
        if not any(cell.strip() for cell in row):
            continue
        if has_header:
            rule = {
                key: row[i].strip()
                for i, key in enumerate(header_keys)
                if key and i < len(row)
            }
        else:
            rule = {
                key: row[i].strip()
                for i, key in enumerate(DEFAULT_ORDER)
                if i < len(row)
            }

        keyword = rule.get("keyword", rule.get("trigger_condition", "")).strip()
        reply = rule.get("reply_template", "").strip()
        if not keyword or not reply:
            continue

        try:
            priority = int(rule.get("priority", 0) or 0)
        except (ValueError, TypeError):
            priority = 0

        rules.append({
            "keyword": keyword,
            "match_type": map_match_type(rule.get("match_type", "")),
            "reply_template": reply,
            "category": rule.get("category", ""),
            "target_type": rule.get("target_type", "ALL"),
            "target_names": rule.get("target_names", "[]"),
            "priority": priority,
            "enabled": parse_enabled(rule.get("enabled")),
        })
    return rules


def parse_excel_content(file_bytes):
    """解析 Excel (.xlsx) 内容为规则列表"""
    import openpyxl
    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    all_rules = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([
                str(cell).strip() if cell is not None else ""
                for cell in row
            ])
        rows = [r for r in rows if any(c for c in r)]
        if not rows:
            continue
        csv_text = "\n".join([",".join(r) for r in rows])
        sheet_rules = parse_csv_content(csv_text)
        all_rules.extend(sheet_rules)
    wb.close()
    return all_rules


@app.route("/admin/login", methods=["GET", "POST"])
def login():
    if session.get("admin_phone"):
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        phone = request.form.get("phone", "").strip()
        password = request.form.get("password", "")
        # Simple rate limiting via session
        now = time.time()
        attempts = session.get('_login_attempts', [])
        attempts = [t for t in attempts if now - t < 300]  # 5-minute window
        if len(attempts) >= 5:
            error = "登录尝试次数过多，请5分钟后再试"
            return render_template("login.html", error=error)
        attempts.append(now)
        session['_login_attempts'] = attempts
        try:
            resp = api_post("/api/admin/login", {"phone": phone, "password": password})
            if resp.status_code == 200:
                result = resp.json()
                if not result.get("is_admin"):
                    error = "需要管理员权限"
                    return render_template("login.html", error=error)
                session["admin_phone"] = result.get("phone", phone)
                session["admin_token"] = result.get("token", "")
                session["is_admin"] = 1
                session.permanent = True
                session['_login_attempts'] = []
                return redirect(url_for("dashboard"))
            else:
                error = _safe_api_error(resp)
        except http_requests.exceptions.RequestException:
            error = "无法连接 API 服务"
        return render_template("login.html", error=error)
    return render_template("login.html", error=None)


@app.route("/admin/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/admin")
def admin_index():
    return redirect(url_for("dashboard"))


@app.route("/admin/dashboard")
@login_required
def dashboard():
    token = session.get("admin_token", "")
    stats = {}
    recent_tenants = []
    try:
        resp = api_get("/api/admin/stats", token)
        if resp.status_code == 200:
            stats = resp.json()
    except Exception:
        pass
    try:
        resp = api_get("/api/admin/recent-tenants", token)
        if resp.status_code == 200:
            recent_tenants = resp.json()
    except Exception:
        pass
    return render_template("dashboard.html", stats=stats, recent_tenants=recent_tenants,
                           admin_phone=session.get("admin_phone"))


@app.route("/admin/tenants")
@login_required
def tenants():
    token = session.get("admin_token", "")
    page = request.args.get("page", 1, type=int)
    search = request.args.get("search", "").strip()
    status = request.args.get("status", "all").strip()
    tenants_data = {"items": [], "total": 0, "page": page, "page_size": 20}
    try:
        params = {"page": page, "page_size": 20}
        if search:
            params["search"] = search
        if status and status != "all":
            params["status"] = status
        query = "/api/admin/tenants?" + urllib.parse.urlencode(params)
        resp = api_get(query, token)
        if resp.status_code == 200:
            tenants_data = resp.json()
    except Exception:
        pass
    return render_template(
        "tenants.html",
        tenants=tenants_data.get("items", []),
        total=tenants_data.get("total", 0),
        page=page,
        page_size=20,
        search=search,
        status=status,
        admin_phone=session.get("admin_phone"),
    )


@app.route("/admin/tenants/<tenant_id>", methods=["GET", "POST"])
@login_required
def tenant_detail(tenant_id):
    token = session.get("admin_token", "")
    error = None
    success = None

    if request.method == "POST":
        action = request.form.get("action")
        if action == "toggle_status":
            is_active = request.form.get("is_active")
            if is_active is not None:
                try:
                    resp = api_put(f"/api/admin/tenants/{tenant_id}", token, {"is_active": int(is_active)})
                    if resp.status_code == 200:
                        flash("状态已更新", "success")
                    else:
                        flash(_safe_api_error(resp), "error")
                except Exception as e:
                    flash(f"网络错误: {e}", "error")
            return redirect(url_for("tenant_detail", tenant_id=tenant_id))
        elif action == "save_default_model":
            model_data = {
                "name": request.form.get("name", "").strip(),
                "model_type": request.form.get("model_type", "OPENAI").strip(),
                "model": request.form.get("model", "").strip(),
                "api_key": request.form.get("api_key", "").strip(),
                "api_endpoint": request.form.get("api_endpoint", "").strip(),
                "temperature": _safe_float(request.form.get("temperature"), 0.7),
                "max_tokens": _safe_int(request.form.get("max_tokens"), 2000),
                "enabled": 1 if request.form.get("enabled") else 0,
            }
            try:
                resp = api_post(f"/api/admin/tenants/{tenant_id}/default-model", model_data, token=token)
                if resp.status_code in (200, 201):
                    success = "默认模型配置已保存"
                else:
                    error = _safe_api_error(resp)
            except Exception as e:
                error = f"网络错误: {e}"
            return render_template("tenant_detail.html", tenant=_get_tenant(tenant_id, token),
                                   default_model=_get_default_model(tenant_id, token),
                                   admin_phone=session.get("admin_phone"), error=error, success=success)
        elif action == "delete_default_model":
            try:
                resp = api_delete(f"/api/admin/tenants/{tenant_id}/default-model", token)
                if resp.status_code == 200:
                    success = "默认模型配置已删除"
                else:
                    error = resp.json().get("error", "删除失败")
            except Exception as e:
                error = f"网络错误: {e}"
            return render_template("tenant_detail.html", tenant=_get_tenant(tenant_id, token),
                                   default_model=None,
                                   admin_phone=session.get("admin_phone"), error=error, success=success)

    tenant = _get_tenant(tenant_id, token)
    default_model = _get_default_model(tenant_id, token)
    return render_template("tenant_detail.html", tenant=tenant, default_model=default_model,
                           admin_phone=session.get("admin_phone"), error=error, success=success)


def _get_tenant(tenant_id, token):
    try:
        resp = api_get(f"/api/admin/tenants/{tenant_id}", token)
        if resp.status_code == 200:
            return resp.json()
    except Exception:
        pass
    return {}


def _get_default_model(tenant_id, token):
    try:
        resp = api_get(f"/api/admin/tenants/{tenant_id}/default-model", token)
        if resp.status_code == 200:
            return resp.json()
    except Exception:
        pass
    return None


@app.route("/admin/default-model", methods=["GET", "POST"])
@login_required
def default_model():
    token = session.get("admin_token", "")
    error = None
    success = None

    if request.method == "POST":
        action = request.form.get("action")
        if action == "apply_to_tenant":
            tenant_id = request.form.get("tenant_id", "").strip()
            if not tenant_id:
                error = "请输入租户 ID"
            else:
                # 获取当前全局默认配置
                current = None
                try:
                    resp = api_get("/api/admin/tenants/_global/default-model", token)
                    if resp.status_code == 200:
                        current = resp.json()
                except Exception:
                    pass
                if not current:
                    error = "请先保存全局默认模型配置"
                else:
                    model_data = {
                        "name": current.get("name", ""),
                        "model_type": current.get("model_type", "OPENAI"),
                        "model": current.get("model", ""),
                        "api_key": current.get("api_key", ""),
                        "api_endpoint": current.get("api_endpoint", ""),
                        "temperature": current.get("temperature", 0.7),
                        "max_tokens": current.get("max_tokens", 2000),
                        "enabled": 1,
                    }
                    try:
                        resp = api_post(f"/api/admin/tenants/{tenant_id}/default-model", model_data, token=token)
                        if resp.status_code in (200, 201):
                            success = f"已应用到租户 {tenant_id[:12]}..."
                        else:
                            error = resp.json().get("error", "应用失败")
                    except Exception as e:
                        error = f"网络错误: {e}"
        else:
            # 保存全局默认（tenant_id = _global）
            model_data = {
                "name": request.form.get("name", "").strip(),
                "model_type": request.form.get("model_type", "OPENAI").strip(),
                "model": request.form.get("model", "").strip(),
                "api_key": request.form.get("api_key", "").strip(),
                "api_endpoint": request.form.get("api_endpoint", "").strip(),
                "temperature": _safe_float(request.form.get("temperature"), 0.7),
                "max_tokens": _safe_int(request.form.get("max_tokens"), 2000),
                "enabled": 1 if request.form.get("enabled") else 0,
            }
            try:
                resp = api_post("/api/admin/tenants/_global/default-model", model_data, token=token)
                if resp.status_code in (200, 201):
                    success = "全局默认模型已保存"
                else:
                    error = resp.json().get("error", "保存失败")
            except Exception as e:
                error = f"网络错误: {e}"

    # 获取当前全局默认
    current_model = None
    try:
        resp = api_get("/api/admin/tenants/_global/default-model", token)
        if resp.status_code == 200:
            current_model = resp.json()
    except Exception:
        pass

    return render_template("default_model.html", model=current_model,
                           admin_phone=session.get("admin_phone"), error=error, success=success)


@app.route("/admin/tenants/<tenant_id>/rules")
@login_required
def admin_tenant_rules(tenant_id):
    """租户知识库规则管理页面"""
    token = session.get("admin_token", "")
    tenant = _get_tenant(tenant_id, token)
    rules = []
    try:
        resp = api_get(f"/api/admin/tenants/{tenant_id}/rules", token)
        if resp.status_code == 200:
            rules = resp.json()
    except Exception:
        pass
    return render_template(
        "tenant_rules.html",
        tenant=tenant,
        rules=rules,
        admin_phone=session.get("admin_phone"),
    )


@app.route("/admin/tenants/<tenant_id>/rules/add", methods=["POST"])
@login_required
def admin_tenant_rule_add(tenant_id):
    """管理员为租户新增规则"""
    token = session.get("admin_token", "")
    data = {
        "keyword": request.form.get("keyword", "").strip(),
        "match_type": request.form.get("match_type", "CONTAINS").strip(),
        "reply_template": request.form.get("reply_template", "").strip(),
        "category": request.form.get("category", "").strip(),
        "target_type": request.form.get("target_type", "ALL").strip(),
        "target_names": request.form.get("target_names", "[]").strip(),
        "priority": _safe_int(request.form.get("priority"), 0),
        "enabled": 1 if request.form.get("enabled") else 0,
    }
    try:
        resp = api_post(f"/api/admin/tenants/{tenant_id}/rules", data, token, timeout=30)
        if resp.status_code in (200, 201):
            flash("规则添加成功", "success")
        else:
            flash(resp.json().get("error", "添加失败"), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_tenant_rules", tenant_id=tenant_id))


@app.route("/admin/tenants/<tenant_id>/rules/<int:rule_id>/edit", methods=["POST"])
@login_required
def admin_tenant_rule_edit(tenant_id, rule_id):
    """管理员编辑租户规则"""
    token = session.get("admin_token", "")
    data = {
        "keyword": request.form.get("keyword", "").strip(),
        "match_type": request.form.get("match_type", "CONTAINS").strip(),
        "reply_template": request.form.get("reply_template", "").strip(),
        "category": request.form.get("category", "").strip(),
        "target_type": request.form.get("target_type", "ALL").strip(),
        "target_names": request.form.get("target_names", "[]").strip(),
        "priority": _safe_int(request.form.get("priority"), 0),
        "enabled": 1 if request.form.get("enabled") else 0,
    }
    try:
        resp = api_put(f"/api/admin/tenants/{tenant_id}/rules/{rule_id}", token, data)
        if resp.status_code == 200:
            flash("规则更新成功", "success")
        else:
            flash(resp.json().get("error", "更新失败"), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_tenant_rules", tenant_id=tenant_id))


@app.route("/admin/tenants/<tenant_id>/rules/<int:rule_id>/delete", methods=["POST"])
@login_required
def admin_tenant_rule_delete(tenant_id, rule_id):
    """管理员删除租户规则"""
    token = session.get("admin_token", "")
    try:
        resp = api_delete(f"/api/admin/tenants/{tenant_id}/rules/{rule_id}", token)
        if resp.status_code == 200:
            flash("规则已删除", "success")
        else:
            flash(resp.json().get("error", "删除失败"), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_tenant_rules", tenant_id=tenant_id))


@app.route("/admin/tenants/<tenant_id>/rules/batch", methods=["POST"])
@login_required
def admin_tenant_rules_batch(tenant_id):
    """批量导入规则，支持文件上传（JSON/CSV/Excel）和 JSON 文本粘贴"""
    token = session.get("admin_token", "")
    import_mode = request.form.get("import_mode", "override")
    if import_mode not in ("override", "append"):
        import_mode = "override"
    rules = []
    source_desc = ""

    # 优先处理文件上传
    uploaded_file = request.files.get("import_file")
    if uploaded_file and uploaded_file.filename:
        filename = uploaded_file.filename.lower()
        # Validate extension
        ALLOWED_EXTENSIONS = {".json", ".csv", ".xlsx"}
        if not any(filename.endswith(ext) for ext in ALLOWED_EXTENSIONS):
            flash("不支持的文件格式，请上传 .json、.csv 或 .xlsx 文件", "error")
            return redirect(url_for("admin_tenant_rules", tenant_id=tenant_id))
        file_bytes = uploaded_file.read()
        # Validate file size (already limited by MAX_CONTENT_LENGTH, but double-check)
        if len(file_bytes) > 10 * 1024 * 1024:
            flash("文件大小超过 10MB 限制", "error")
            return redirect(url_for("admin_tenant_rules", tenant_id=tenant_id))
        try:
            if filename.endswith(".csv"):
                rules = parse_csv_content(file_bytes.decode("utf-8-sig"))
                source_desc = "CSV 文件"
            elif filename.endswith(".xlsx"):
                rules = parse_excel_content(file_bytes)
                source_desc = "Excel 文件"
            elif filename.endswith(".json"):
                rules = parse_json_content(file_bytes.decode("utf-8-sig"))
                source_desc = "JSON 文件"

        except Exception as e:
            flash(f"文件解析失败: {e}", "error")
            return redirect(url_for("admin_tenant_rules", tenant_id=tenant_id))
    else:
        # 回退到文本粘贴
        import_data = request.form.get("import_data", "").strip()
        if not import_data:
            flash("请上传文件或粘贴 JSON 数据", "error")
            return redirect(url_for("admin_tenant_rules", tenant_id=tenant_id))
        try:
            rules = parse_json_content(import_data)
            source_desc = "JSON 文本"
        except json.JSONDecodeError as e:
            flash(f"JSON 格式错误: {e}", "error")
            return redirect(url_for("admin_tenant_rules", tenant_id=tenant_id))

    if not rules:
        flash(f"{source_desc} 中没有可导入的规则", "error")
        return redirect(url_for("admin_tenant_rules", tenant_id=tenant_id))

    if len(rules) > 1000:
        flash(f"规则数量超过上限（1000条），当前 {len(rules)} 条", "error")
        return redirect(url_for("admin_tenant_rules", tenant_id=tenant_id))

    try:
        resp = api_post(
            f"/api/admin/tenants/{tenant_id}/rules/batch",
            {"rules": rules, "mode": import_mode},
            token,
            timeout=30,
        )
        if resp.status_code == 200:
            try:
                result = resp.json()
                imported = result.get("imported", len(rules))
                total = result.get("total", imported)
            except Exception:
                imported = len(rules)
                total = imported
            mode_desc = "覆盖导入" if import_mode == "override" else "追加导入"
            flash(
                f"{mode_desc}成功：从 {source_desc} 导入 {imported} 条规则，当前共 {total} 条",
                "success",
            )
        else:
            flash(_safe_api_error(resp), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_tenant_rules", tenant_id=tenant_id))


@app.route("/admin/tenants/<tenant_id>/blacklist")
@login_required
def admin_tenant_blacklist(tenant_id):
    """租户黑名单管理页面"""
    token = session.get("admin_token", "")
    tenant = _get_tenant(tenant_id, token)
    blacklists = []
    try:
        resp = api_get(f"/api/admin/tenants/{tenant_id}/blacklist", token)
        if resp.status_code == 200:
            blacklists = resp.json()
    except Exception:
        pass
    return render_template(
        "tenant_blacklist.html",
        tenant=tenant,
        blacklists=blacklists,
        admin_phone=session.get("admin_phone"),
    )


@app.route("/admin/tenants/<tenant_id>/blacklist/add", methods=["POST"])
@login_required
def admin_tenant_blacklist_add(tenant_id):
    """管理员为租户新增黑名单条目"""
    token = session.get("admin_token", "")
    data = {
        "type": request.form.get("type", "KEYWORD").strip(),
        "value": request.form.get("value", "").strip(),
        "description": request.form.get("description", "").strip(),
        "package_name": request.form.get("package_name", "").strip() or None,
        "is_enabled": 1 if request.form.get("is_enabled") else 0,
    }
    try:
        resp = api_post(f"/api/admin/tenants/{tenant_id}/blacklist", data, token, timeout=30)
        if resp.status_code in (200, 201):
            flash("黑名单添加成功", "success")
        else:
            flash(resp.json().get("error", "添加失败"), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_tenant_blacklist", tenant_id=tenant_id))


@app.route("/admin/tenants/<tenant_id>/blacklist/<int:blacklist_id>/edit", methods=["POST"])
@login_required
def admin_tenant_blacklist_edit(tenant_id, blacklist_id):
    """管理员编辑租户黑名单条目"""
    token = session.get("admin_token", "")
    data = {
        "type": request.form.get("type", "KEYWORD").strip(),
        "value": request.form.get("value", "").strip(),
        "description": request.form.get("description", "").strip(),
        "package_name": request.form.get("package_name", "").strip() or None,
        "is_enabled": 1 if request.form.get("is_enabled") else 0,
    }
    try:
        resp = api_put(f"/api/admin/tenants/{tenant_id}/blacklist/{blacklist_id}", token, data)
        if resp.status_code == 200:
            flash("黑名单更新成功", "success")
        else:
            flash(resp.json().get("error", "更新失败"), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_tenant_blacklist", tenant_id=tenant_id))


@app.route("/admin/tenants/<tenant_id>/blacklist/<int:blacklist_id>/delete", methods=["POST"])
@login_required
def admin_tenant_blacklist_delete(tenant_id, blacklist_id):
    """管理员删除租户黑名单条目"""
    token = session.get("admin_token", "")
    try:
        resp = api_delete(f"/api/admin/tenants/{tenant_id}/blacklist/{blacklist_id}", token)
        if resp.status_code == 200:
            flash("黑名单已删除", "success")
        else:
            flash(resp.json().get("error", "删除失败"), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_tenant_blacklist", tenant_id=tenant_id))


@app.route("/admin/tenants/<tenant_id>/blacklist/clear", methods=["POST"])
@login_required
def admin_tenant_blacklist_clear(tenant_id):
    """管理员清空租户所有黑名单"""
    token = session.get("admin_token", "")
    try:
        resp = api_post(f"/api/admin/tenants/{tenant_id}/blacklist/clear", {}, token, timeout=30)
        if resp.status_code == 200:
            flash("已清空所有黑名单", "success")
        else:
            flash(resp.json().get("error", "清空失败"), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_tenant_blacklist", tenant_id=tenant_id))


@app.route("/admin/tenants/<tenant_id>/history")
@login_required
def admin_tenant_history(tenant_id):
    """租户回复历史查看页面"""
    token = session.get("admin_token", "")
    tenant = _get_tenant(tenant_id, token)
    page = request.args.get("page", 1, type=int)
    page_size = _clamp_page_size(request.args.get("page_size", 20))
    source = request.args.get("source", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()

    history_data = {"items": [], "total": 0, "page": page, "page_size": page_size}
    try:
        params = {"page": page, "page_size": page_size}
        if source:
            params["source"] = source
        if date_from:
            params["date_from"] = date_from
        if date_to:
            params["date_to"] = date_to
        query = f"/api/admin/tenants/{tenant_id}/history?" + urllib.parse.urlencode(params)
        resp = api_get(query, token)
        if resp.status_code == 200:
            history_data = resp.json()
    except Exception:
        pass
    return render_template(
        "tenant_history.html",
        tenant=tenant,
        items=history_data.get("items", []),
        total=history_data.get("total", 0),
        page=history_data.get("page", page),
        page_size=history_data.get("page_size", page_size),
        source=source,
        date_from=date_from,
        date_to=date_to,
        admin_phone=session.get("admin_phone"),
    )


@app.route("/admin/tenants/<tenant_id>/models")
@login_required
def admin_tenant_models(tenant_id):
    """租户模型配置管理页面"""
    token = session.get("admin_token", "")
    tenant = _get_tenant(tenant_id, token)
    models = []
    try:
        resp = api_get(f"/api/admin/tenants/{tenant_id}/models", token)
        if resp.status_code == 200:
            models = resp.json()
    except Exception:
        pass
    return render_template(
        "tenant_models.html",
        tenant=tenant,
        models=models,
        admin_phone=session.get("admin_phone"),
    )


@app.route("/admin/tenants/<tenant_id>/models/add", methods=["POST"])
@login_required
def admin_tenant_model_add(tenant_id):
    """管理员为租户新增模型配置"""
    token = session.get("admin_token", "")
    data = {
        "name": request.form.get("name", "").strip(),
        "model_type": request.form.get("model_type", "OPENAI").strip(),
        "model": request.form.get("model", "").strip(),
        "api_key": request.form.get("api_key", "").strip(),
        "api_endpoint": request.form.get("api_endpoint", "").strip(),
        "temperature": _safe_float(request.form.get("temperature"), 0.7),
        "max_tokens": _safe_int(request.form.get("max_tokens"), 2000),
        "is_default": 1 if request.form.get("is_default") else 0,
        "enabled": 1 if request.form.get("enabled") else 0,
    }
    try:
        resp = api_post(f"/api/admin/tenants/{tenant_id}/models", data, token, timeout=30)
        if resp.status_code in (200, 201):
            flash("模型配置添加成功", "success")
        else:
            flash(resp.json().get("error", "添加失败"), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_tenant_models", tenant_id=tenant_id))


@app.route("/admin/tenants/<tenant_id>/models/<int:model_id>/edit", methods=["POST"])
@login_required
def admin_tenant_model_edit(tenant_id, model_id):
    """管理员编辑租户模型配置"""
    token = session.get("admin_token", "")
    data = {
        "name": request.form.get("name", "").strip(),
        "model_type": request.form.get("model_type", "OPENAI").strip(),
        "model": request.form.get("model", "").strip(),
        "api_key": request.form.get("api_key", "").strip(),
        "api_endpoint": request.form.get("api_endpoint", "").strip(),
        "temperature": _safe_float(request.form.get("temperature"), 0.7),
        "max_tokens": _safe_int(request.form.get("max_tokens"), 2000),
        "is_default": 1 if request.form.get("is_default") else 0,
        "enabled": 1 if request.form.get("enabled") else 0,
    }
    try:
        resp = api_put(f"/api/admin/tenants/{tenant_id}/models/{model_id}", token, data)
        if resp.status_code == 200:
            flash("模型配置更新成功", "success")
        else:
            flash(resp.json().get("error", "更新失败"), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_tenant_models", tenant_id=tenant_id))


@app.route("/admin/tenants/<tenant_id>/models/<int:model_id>/delete", methods=["POST"])
@login_required
def admin_tenant_model_delete(tenant_id, model_id):
    """管理员删除租户模型配置"""
    token = session.get("admin_token", "")
    try:
        resp = api_delete(f"/api/admin/tenants/{tenant_id}/models/{model_id}", token)
        if resp.status_code == 200:
            flash("模型配置已删除", "success")
        else:
            flash(resp.json().get("error", "删除失败"), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_tenant_models", tenant_id=tenant_id))


@app.route("/admin/tenants/<tenant_id>/feedback")
@login_required
def admin_tenant_feedback(tenant_id):
    """租户用户反馈查看页面"""
    token = session.get("admin_token", "")
    tenant = _get_tenant(tenant_id, token)
    page = request.args.get("page", 1, type=int)
    page_size = _clamp_page_size(request.args.get("page_size", 20))
    action = request.args.get("action", "").strip()

    feedback_data = {"items": [], "total": 0, "page": page, "page_size": page_size}
    try:
        params = {"page": page, "page_size": page_size}
        if action:
            params["action"] = action
        query = f"/api/admin/tenants/{tenant_id}/feedback?" + urllib.parse.urlencode(params)
        resp = api_get(query, token)
        if resp.status_code == 200:
            feedback_data = resp.json()
    except Exception:
        pass
    return render_template(
        "tenant_feedback.html",
        tenant=tenant,
        items=feedback_data.get("items", []),
        total=feedback_data.get("total", 0),
        page=feedback_data.get("page", page),
        page_size=feedback_data.get("page_size", page_size),
        action=action,
        admin_phone=session.get("admin_phone"),
    )


@app.route("/admin/tenants/<tenant_id>/metrics")
@login_required
def admin_tenant_metrics(tenant_id):
    """租户优化指标查看页面"""
    token = session.get("admin_token", "")
    tenant = _get_tenant(tenant_id, token)
    days = request.args.get("days", 7, type=int)
    days = max(1, min(days, 365))
    page = request.args.get("page", 1, type=int)
    page_size = _clamp_page_size(request.args.get("page_size", 30))

    summary = {}
    metrics_data = {"items": [], "total": 0, "page": page, "page_size": page_size}
    try:
        resp = api_get(f"/api/admin/tenants/{tenant_id}/metrics/summary?days={int(days)}", token)
        if resp.status_code == 200:
            summary = resp.json()
    except Exception:
        pass
    try:
        resp = api_get(f"/api/admin/tenants/{tenant_id}/metrics?days={days}&page={page}&page_size={page_size}", token)
        if resp.status_code == 200:
            metrics_data = resp.json()
    except Exception:
        pass
    return render_template(
        "tenant_metrics.html",
        tenant=tenant,
        summary=summary,
        items=metrics_data.get("items", []),
        total=metrics_data.get("total", 0),
        page=metrics_data.get("page", page),
        page_size=metrics_data.get("page_size", page_size),
        days=days,
        admin_phone=session.get("admin_phone"),
    )


@app.route("/admin/audit-log")
@login_required
def admin_audit_log():
    """平台操作日志页面"""
    token = session.get("admin_token", "")
    page = request.args.get("page", 1, type=int)
    action = request.args.get("action", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()

    audit_data = {"items": [], "total": 0, "page": page, "page_size": 20}
    try:
        params = {"page": page, "page_size": 20}
        if action:
            params["action"] = action
        if date_from:
            params["date_from"] = date_from
        if date_to:
            params["date_to"] = date_to
        query = "/api/admin/audit-log?" + urllib.parse.urlencode(params)
        resp = api_get(query, token)
        if resp.status_code == 200:
            audit_data = resp.json()
    except Exception:
        pass
    return render_template(
        "audit_log.html",
        items=audit_data.get("items", []),
        total=audit_data.get("total", 0),
        page=audit_data.get("page", page),
        page_size=audit_data.get("page_size", 20),
        action=action,
        date_from=date_from,
        date_to=date_to,
        admin_phone=session.get("admin_phone"),
    )


@app.route("/admin/admins")
@login_required
def admin_admins():
    """管理员账户管理页面"""
    token = session.get("admin_token", "")
    admins = []
    try:
        resp = api_get("/api/admin/admins", token)
        if resp.status_code == 200:
            admins = resp.json()
    except Exception:
        pass
    return render_template(
        "admins.html",
        admins=admins,
        admin_phone=session.get("admin_phone"),
    )


@app.route("/admin/admins/add", methods=["POST"])
@login_required
def admin_admins_add():
    """创建管理员"""
    token = session.get("admin_token", "")
    data = {
        "phone": request.form.get("phone", "").strip(),
        "password": request.form.get("password", ""),
        "is_active": 1 if request.form.get("is_active") else 0,
    }
    try:
        resp = api_post("/api/admin/admins", data, token=token)
        if resp.status_code in (200, 201):
            flash("管理员创建成功", "success")
        else:
            flash(resp.json().get("error", "创建失败"), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_admins"))


@app.route("/admin/admins/<int:admin_id>/edit", methods=["POST"])
@login_required
def admin_admins_edit(admin_id):
    """编辑管理员"""
    token = session.get("admin_token", "")
    data = {}
    is_active = request.form.get("is_active")
    if is_active is not None:
        data["is_active"] = 1 if is_active else 0
    password = request.form.get("password", "")
    confirm_password = request.form.get("confirm_password", "")
    if password:
        if password != confirm_password:
            flash("两次密码不一致", "error")
            return redirect(url_for("admin_admins"))
        if len(password) < 6:
            flash("密码至少6位", "error")
            return redirect(url_for("admin_admins"))
        data["password"] = password
    try:
        resp = api_put(f"/api/admin/admins/{admin_id}", token, data)
        if resp.status_code == 200:
            flash("管理员信息已更新", "success")
        else:
            flash(resp.json().get("error", "更新失败"), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_admins"))


@app.route("/admin/admins/<int:admin_id>/delete", methods=["POST"])
@login_required
def admin_admins_delete(admin_id):
    """删除管理员"""
    token = session.get("admin_token", "")
    try:
        resp = api_delete(f"/api/admin/admins/{admin_id}", token)
        if resp.status_code == 200:
            flash("管理员已删除", "success")
        else:
            flash(resp.json().get("error", "删除失败"), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_admins"))


@app.route("/admin/profile", methods=["GET", "POST"])
@login_required
def profile():
    token = session.get("admin_token", "")
    if request.method == "POST":
        old_password = request.form.get("old_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")
        if not old_password or not new_password:
            return render_template("profile.html", admin_phone=session.get("admin_phone"), error="请填写完整", success=None)
        if new_password != confirm_password:
            return render_template("profile.html", admin_phone=session.get("admin_phone"), error="两次新密码不一致", success=None)
        if len(new_password) < 6:
            return render_template("profile.html", admin_phone=session.get("admin_phone"), error="新密码至少6位", success=None)
        try:
            resp = api_post(
                "/api/auth/change_password",
                {"old_password": old_password, "new_password": new_password},
                token=token,
            )
            if resp.status_code == 200:
                return render_template("profile.html", admin_phone=session.get("admin_phone"), error=None, success="密码修改成功")
            else:
                error = resp.json().get("error", "修改失败")
                return render_template("profile.html", admin_phone=session.get("admin_phone"), error=error, success=None)
        except Exception as e:
            return render_template("profile.html", admin_phone=session.get("admin_phone"), error=f"网络错误: {e}", success=None)
    return render_template("profile.html", admin_phone=session.get("admin_phone"), error=None, success=None)


import re

# ========== Smart Routing Admin Pages ==========


@app.route("/admin/tenants/<tenant_id>/routing")
@login_required
def admin_tenant_routing(tenant_id):
    """智能路由管理页面：客服状态 + 会话列表 + 路由配置"""
    token = session.get("admin_token", "")
    tenant = _get_tenant(tenant_id, token)

    # Fetch agents
    agents = []
    try:
        resp = api_get(f"/api/admin/tenants/{tenant_id}/agents", token)
        if resp.status_code == 200:
            agents = resp.json().get("agents", [])
    except Exception:
        pass

    # Fetch sessions
    sessions = []
    sessions_total = 0
    try:
        resp = api_get(f"/api/admin/tenants/{tenant_id}/sessions", token)
        if resp.status_code == 200:
            result = resp.json()
            sessions = result.get("sessions", [])
            sessions_total = result.get("total", 0)
    except Exception:
        pass

    # Fetch routing config
    routing_config = None
    try:
        resp = api_get(f"/api/admin/tenants/{tenant_id}/routing/config", token)
        if resp.status_code == 200:
            routing_config = resp.json()
    except Exception:
        pass

    return render_template(
        "tenant_routing.html",
        tenant=tenant,
        agents=agents,
        sessions=sessions,
        sessions_total=sessions_total,
        routing_config=routing_config,
        admin_phone=session.get("admin_phone"),
    )


@app.route("/admin/tenants/<tenant_id>/routing/agent/add", methods=["POST"])
@login_required
def admin_tenant_routing_agent_add(tenant_id):
    """添加客服并设置状态"""
    token = session.get("admin_token", "")
    agent_phone = request.form.get("agent_phone", "").strip()
    agent_name = request.form.get("agent_name", "").strip()
    status = request.form.get("status", "online").strip()
    max_concurrent = request.form.get("max_concurrent", "5").strip()

    if not agent_phone:
        flash("客服手机号不能为空", "error")
        return redirect(url_for("admin_tenant_routing", tenant_id=tenant_id))

    try:
        max_concurrent = int(max_concurrent)
    except ValueError:
        max_concurrent = 5

    data = {
        "agent_phone": agent_phone,
        "agent_name": agent_name,
        "status": status,
        "max_concurrent": max_concurrent,
    }
    try:
        resp = api_post(f"/api/agent/status", data, token, timeout=10)
        if resp.status_code in (200, 201):
            flash(f"客服 {agent_phone} 添加成功", "success")
        else:
            flash(_safe_api_error(resp), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_tenant_routing", tenant_id=tenant_id))


@app.route("/admin/tenants/<tenant_id>/routing/agent/<agent_phone>/status", methods=["POST"])
@login_required
def admin_tenant_routing_agent_status(tenant_id, agent_phone):
    """更新客服状态"""
    token = session.get("admin_token", "")
    status = request.form.get("status", "").strip()
    if status not in ("online", "offline", "busy", "away"):
        flash("无效的状态值", "error")
        return redirect(url_for("admin_tenant_routing", tenant_id=tenant_id))

    data = {"agent_phone": agent_phone, "status": status}
    try:
        resp = api_post(f"/api/agent/status", data, token, timeout=10)
        if resp.status_code == 200:
            flash(f"客服 {agent_phone} 状态已更新为 {status}", "success")
        else:
            flash(_safe_api_error(resp), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_tenant_routing", tenant_id=tenant_id))


@app.route("/admin/tenants/<tenant_id>/routing/agent/<agent_phone>/skills", methods=["POST"])
@login_required
def admin_tenant_routing_agent_skills(tenant_id, agent_phone):
    """设置客服技能"""
    token = session.get("admin_token", "")
    skills_raw = request.form.get("skills", "").strip()
    skills = []
    if skills_raw:
        for line in skills_raw.split("\n"):
            line = line.strip()
            if not line:
                continue
            parts = line.split(",")
            tag = parts[0].strip()
            prof = 5
            if len(parts) > 1:
                try:
                    prof = int(parts[1].strip())
                except ValueError:
                    prof = 5
            skills.append({"skill_tag": tag, "proficiency": prof})

    data = {"agent_phone": agent_phone, "skills": skills}
    try:
        resp = api_post(f"/api/agent/skills", data, token, timeout=10)
        if resp.status_code == 200:
            flash(f"客服 {agent_phone} 技能已更新", "success")
        else:
            flash(_safe_api_error(resp), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_tenant_routing", tenant_id=tenant_id))


@app.route("/admin/tenants/<tenant_id>/routing/config", methods=["POST"])
@login_required
def admin_tenant_routing_config(tenant_id):
    """更新路由配置"""
    token = session.get("admin_token", "")
    strategy = request.form.get("strategy", "skill_first").strip()
    fallback_to_ai = 1 if request.form.get("fallback_to_ai") else 0
    max_queue_size = request.form.get("max_queue_size", "50").strip()
    timeout_seconds = request.form.get("timeout_seconds", "300").strip()

    try:
        max_queue_size = int(max_queue_size)
    except ValueError:
        max_queue_size = 50
    try:
        timeout_seconds = int(timeout_seconds)
    except ValueError:
        timeout_seconds = 300

    data = {
        "strategy": strategy,
        "fallback_to_ai": fallback_to_ai,
        "max_queue_size": max_queue_size,
        "timeout_seconds": timeout_seconds,
    }
    try:
        resp = api_post(f"/api/routing/config", data, token, timeout=10)
        if resp.status_code == 200:
            flash("路由配置已更新", "success")
        else:
            flash(_safe_api_error(resp), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_tenant_routing", tenant_id=tenant_id))


@app.route("/admin/tenants/<tenant_id>/routing/session/<int:session_id>/close", methods=["POST"])
@login_required
def admin_tenant_routing_session_close(tenant_id, session_id):
    """关闭会话"""
    token = session.get("admin_token", "")
    agent_phone = request.form.get("agent_phone", "").strip()
    if not agent_phone:
        flash("需要指定客服手机号", "error")
        return redirect(url_for("admin_tenant_routing", tenant_id=tenant_id))

    try:
        resp = api_post(f"/api/conversation/{session_id}/close",
                        {"agent_phone": agent_phone}, token, timeout=10)
        if resp.status_code == 200:
            flash("会话已关闭", "success")
        else:
            flash(_safe_api_error(resp), "error")
    except Exception as e:
        flash(f"网络错误: {e}", "error")
    return redirect(url_for("admin_tenant_routing", tenant_id=tenant_id))


@app.after_request
def _inject_csrf_token(response):
    """Inject CSRF token into all HTML forms and set security headers."""
    # Security headers
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "style-src 'self' 'unsafe-inline'; "
        "script-src 'self' 'unsafe-inline'; "
        "img-src 'self' data:; "
        "font-src 'self'; "
        "frame-ancestors 'none'"
    )
    if response.content_type and "text/html" in response.content_type:
        token = session.get("_csrf_token", "")
        if not token:
            token = _csrf_token()
        hidden = '<input type="hidden" name="_csrf_token" value="' + token + '">'
        html = response.get_data(as_text=True)
        # Inject CSRF token into all POST forms that don't already have it
        html = re.sub(
            r'(<form\s[^>]*method=["\']POST["\'][^>]*>)',
            r'\1' + hidden,
            html,
            flags=re.IGNORECASE
        )
        response.set_data(html)
    return response

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8080, debug=False)
