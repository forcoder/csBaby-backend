"""Tests for admin CSV/Excel parsing functions.

We import parse_csv_content and parse_excel_content by extracting them
from app.py using runpy to handle the module loading properly.
"""
import pytest
import sys
import os
import types
import runpy

# Load app.py as a module so we can access its functions
_admin_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _admin_dir)

# Set env vars before loading
os.environ.setdefault("API_BASE_URL", "http://localhost:5000")
os.environ.setdefault("SESSION_SECRET", "test-secret-key")

# Load utils first (needed by app)
_utils_mod = runpy.run_path(os.path.join(_admin_dir, "utils.py"))
_app_mod = runpy.run_path(os.path.join(_admin_dir, "app.py"))

parse_csv_content = _app_mod["parse_csv_content"]
parse_excel_content = _app_mod["parse_excel_content"]


class TestParseCsvContent:
    def test_basic_csv_with_headers(self):
        content = "keyword,reply_template\n你好,您好\n谢谢,不客气"
        result = parse_csv_content(content)
        assert len(result) == 2
        assert result[0]["keyword"] == "你好"
        assert result[0]["reply_template"] == "您好"
        assert result[1]["keyword"] == "谢谢"

    def test_csv_with_chinese_headers(self):
        content = "关键词,回复内容\n入住,欢迎入住\n退房,祝您生活愉快"
        result = parse_csv_content(content)
        assert len(result) == 2
        assert result[0]["keyword"] == "入住"
        assert result[0]["reply_template"] == "欢迎入住"

    def test_csv_without_recognizable_headers(self):
        """When no recognizable header is found, parser may return empty
        because it cannot determine column mapping from data-only rows."""
        content = "你好,您好\n谢谢,不客气"
        result = parse_csv_content(content)
        # Without recognizable headers, the parser cannot map columns
        assert isinstance(result, list)

    def test_empty_csv(self):
        result = parse_csv_content("")
        assert result == []

    def test_csv_only_empty_lines(self):
        result = parse_csv_content("\n\n\n")
        assert result == []

    def test_csv_with_all_fields(self):
        content = "keyword,match_type,reply_template,category,target_type,target_names,priority,enabled\n你好,CONTAINS,您好,greeting,ALL,[],10,启用"
        result = parse_csv_content(content)
        assert len(result) == 1
        assert result[0]["keyword"] == "你好"
        assert result[0]["match_type"] == "CONTAINS"
        assert result[0]["priority"] == 10
        assert result[0]["enabled"] == 1

    def test_csv_match_type_mapping_中文(self):
        content = "keyword,触发类型,reply_template\n入住,包含匹配,欢迎\n退房,精确匹配,再见"
        result = parse_csv_content(content)
        assert len(result) == 2
        assert result[0]["match_type"] == "CONTAINS"
        assert result[1]["match_type"] == "EXACT"

    def test_csv_match_type_正则匹配(self):
        content = "keyword,match_type,reply_template\n^你好,REGEX,您好"
        result = parse_csv_content(content)
        assert len(result) == 1
        assert result[0]["match_type"] == "REGEX"

    def test_csv_enabled_禁用(self):
        content = "keyword,reply_template,enabled\n你好,您好,禁用"
        result = parse_csv_content(content)
        assert result[0]["enabled"] == 0

    def test_csv_enabled_true_false(self):
        content = "keyword,reply_template,enabled\n你好,您好,true\n再见,bye,false"
        result = parse_csv_content(content)
        assert result[0]["enabled"] == 1
        assert result[1]["enabled"] == 0

    def test_csv_skips_rows_with_empty_keyword(self):
        content = "keyword,reply_template\n,您好\n谢谢,不客气"
        result = parse_csv_content(content)
        assert len(result) == 1
        assert result[0]["keyword"] == "谢谢"

    def test_csv_skips_rows_with_empty_reply(self):
        content = "keyword,reply_template\n你好,\n谢谢,不客气"
        result = parse_csv_content(content)
        assert len(result) == 1
        assert result[0]["keyword"] == "谢谢"

    def test_csv_default_match_type(self):
        content = "keyword,reply_template\n你好,您好"
        result = parse_csv_content(content)
        assert result[0]["match_type"] == "CONTAINS"

    def test_csv_default_enabled(self):
        content = "keyword,reply_template\n你好,您好"
        result = parse_csv_content(content)
        assert result[0]["enabled"] == 1

    def test_csv_default_priority(self):
        content = "keyword,reply_template\n你好,您好"
        result = parse_csv_content(content)
        assert result[0]["priority"] == 0

    def test_csv_invalid_priority_defaults_to_zero(self):
        content = "keyword,reply_template,priority\n你好,您好,abc"
        result = parse_csv_content(content)
        assert result[0]["priority"] == 0

    def test_csv_with_utf8_bom(self):
        content = "﻿keyword,reply_template\n你好,您好"
        result = parse_csv_content(content)
        assert len(result) == 1
        assert result[0]["keyword"] == "你好"

    def test_csv_with_extra_whitespace(self):
        content = "keyword,reply_template\n  你好  ,  您好  "
        result = parse_csv_content(content)
        assert len(result) == 1
        assert result[0]["keyword"] == "你好"
        assert result[0]["reply_template"] == "您好"

    def test_csv_empty_rows_in_middle(self):
        content = "keyword,reply_template\n你好,您好\n\n谢谢,不客气"
        result = parse_csv_content(content)
        assert len(result) == 2

    def test_csv_rule_title_header(self):
        content = "规则标题,回复内容\n入住,欢迎入住"
        result = parse_csv_content(content)
        assert len(result) == 1
        assert result[0]["keyword"] == "入住"

    def test_csv_trigger_condition_header(self):
        content = "触发条件,reply_template\n入住,欢迎入住"
        result = parse_csv_content(content)
        assert len(result) == 1
        assert result[0]["keyword"] == "入住"

    def test_csv_header_case_insensitive(self):
        content = "KEYWORD,REPLY_TEMPLATE\nhello,hi"
        result = parse_csv_content(content)
        assert len(result) == 1
        assert result[0]["keyword"] == "hello"

    def test_csv_dash_in_header(self):
        content = "match-type,reply_template\nEXACT,您好"
        result = parse_csv_content(content)
        # Should not crash; match_type may not be mapped from dash format
        assert isinstance(result, list)


class TestParseExcelContent:
    def test_excel_basic(self):
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["keyword", "reply_template"])
        ws.append(["你好", "您好"])
        ws.append(["谢谢", "不客气"])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        wb.close()

        result = parse_excel_content(buf.read())
        assert len(result) == 2
        assert result[0]["keyword"] == "你好"
        assert result[1]["keyword"] == "谢谢"

    def test_excel_multiple_sheets(self):
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["keyword", "reply_template"])
        ws1.append(["你好", "您好"])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["keyword", "reply_template"])
        ws2.append(["再见", "拜拜"])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        wb.close()

        result = parse_excel_content(buf.read())
        assert len(result) == 2

    def test_excel_empty_workbook(self):
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        wb.close()

        result = parse_excel_content(buf.read())
        assert result == []

    def test_excel_with_chinese_headers(self):
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["关键词", "回复内容"])
        ws.append(["入住", "欢迎入住"])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        wb.close()

        result = parse_excel_content(buf.read())
        assert len(result) == 1
        assert result[0]["keyword"] == "入住"

    def test_excel_empty_sheet_skipped(self):
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Empty"

        ws2 = wb.create_sheet("Data")
        ws2.append(["keyword", "reply_template"])
        ws2.append(["test", "reply"])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        wb.close()

        result = parse_excel_content(buf.read())
        assert len(result) == 1
        assert result[0]["keyword"] == "test"

    def test_excel_with_all_columns(self):
        import openpyxl
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["keyword", "match_type", "reply_template", "category", "target_type", "target_names", "priority", "enabled"])
        ws.append(["你好", "CONTAINS", "您好！", "greeting", "ALL", "[]", 10, "启用"])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        wb.close()

        result = parse_excel_content(buf.read())
        assert len(result) == 1
        assert result[0]["keyword"] == "你好"
        assert result[0]["match_type"] == "CONTAINS"
        assert result[0]["priority"] == 10
        assert result[0]["enabled"] == 1
